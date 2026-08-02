import openpyxl, re, datetime, unicodedata, shutil, subprocess, tempfile, os
from difflib import SequenceMatcher
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import CharacterProperties, Paragraph, ParagraphProperties, RichTextProperties
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart.data_source import StrRef, AxDataSource
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas as _rl_canvas
    from reportlab.lib import colors as _rl_colors
    from reportlab.pdfbase.pdfmetrics import stringWidth as _rl_stringWidth
except ImportError:
    _rl_canvas = None

# Matricules exclus du classement/podium car responsables (pas des collaborateurs
# a classer). A adapter si l'organisation change.
EXCLUDED_FROM_RANKING = {
    "MONTESCOT3",   # Adrien Navaro - responsable
    "PR0911201",    # Maelle Gendre - responsable
}


def get_week_and_employees(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Preparation']
    periode = ws.cell(row=5, column=3).value  # "Du DD/MM/YYYY Au DD/MM/YYYY"
    m = re.search(r'Du (\d{2})/(\d{2})/(\d{4})', periode)
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    dt = datetime.date(y, mo, d)
    iso = dt.isocalendar()
    week = iso[1]

    def _to_float(v):
        try:
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None

    row = 25
    employees = []  # (name, prep_row, articles_row, articles_count, commandes_count)
    while row <= ws.max_row:
        name = ws.cell(row=row, column=2).value
        if isinstance(name, str) and '(' in name and ')' in name:
            prep_row = row + 1
            articles_row = row + 3
            articles_count = _to_float(ws.cell(row=articles_row, column=3).value)
            commandes_count = _to_float(ws.cell(row=prep_row, column=3).value)
            employees.append((name.strip(), prep_row, articles_row, articles_count, commandes_count))
            row += 12
        else:
            row += 1
    return week, employees


def get_periode(path):
    """Renvoie (date_debut, date_fin) de la periode reellement couverte par
    le fichier Preparation (cellule "Du DD/MM/YYYY Au DD/MM/YYYY"). Renvoie
    (None, None) si le format n'est pas reconnu."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Preparation']
    periode = ws.cell(row=5, column=3).value
    m = re.search(r'Du (\d{2})/(\d{2})/(\d{4}) [Aa]u (\d{2})/(\d{2})/(\d{4})', periode or '')
    if not m:
        return None, None
    d1, mo1, y1, d2, mo2, y2 = (int(g) for g in m.groups())
    return datetime.date(y1, mo1, d1), datetime.date(y2, mo2, d2)


# ---------------------------------------------------------------------------
# Extraction des heures "Drive" depuis un planning PDF (format Boulpat/Drive/Bazar)
# ---------------------------------------------------------------------------

_NAME_RE = re.compile(r'^[A-ZÀ-Ÿ\-]+$')
# Certains intitules de rayon contiennent un "/" (ex: "ECOLE/FORMATION") : une
# regex n'autorisant pas ce caractere fait echouer la reconnaissance de toute
# la ligne de tags rayon pour l'employe concerne, et ses heures DRIVE reelles
# sont alors silencieusement comptees comme 0 (aucun tag ne peut alors
# correspondre a "DRIVE" puisque la ligne entiere est rejetee).
_TAG_RE = re.compile(r'^[A-ZÀ-Ÿ\-/]+$')
_HOUR_RE = re.compile(r'^\d{1,2}h\d{2}$')
_REPOS_TOKENS = {"REPOS", "MALADIE", "ACCIDENT", "CP"}
_NOISE_TOKENS = {
    "Pause", "min", "Reconnais", "SIGNATURE", "avoir", "effectué", "les", "horaires",
    "indiqués", "ci-dessus", "pris", "pauses", "indiquées", "sur", "chaque", "semaine",
    ":", "TOTAL", "Nb", "Heure",
}


def _cluster_rows(words, tol=2.5):
    words = sorted(words, key=lambda w: w['top'])
    rows, cur, cur_top = [], [], None
    for w in words:
        if cur_top is None or abs(w['top'] - cur_top) <= tol:
            cur.append(w)
            cur_top = w['top'] if cur_top is None else cur_top
        else:
            rows.append(cur)
            cur, cur_top = [w], w['top']
    if cur:
        rows.append(cur)
    for r in rows:
        r.sort(key=lambda w: w['x0'])
    return rows


def _hour_to_decimal(s):
    m = re.match(r'^(\d{1,2})h(\d{2})$', s)
    h, mi = int(m.group(1)), int(m.group(2))
    return h + mi / 60.0


def _group_tags(tags_row, gap_threshold=8.0):
    toks = sorted([w for w in tags_row if w['text'] not in _NOISE_TOKENS], key=lambda w: w['x0'])
    groups = []
    for w in toks:
        if groups and (w['x0'] - groups[-1]['x1']) <= gap_threshold:
            groups[-1]['text'] += ' ' + w['text']
            groups[-1]['x1'] = w['x1']
        else:
            groups.append({'text': w['text'], 'x0': w['x0'], 'x1': w['x1']})
    return groups


_WEEKDAYS = {"Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"}
_DATE_DM_RE = re.compile(r'^(\d{2})/(\d{2})$')


def _extract_day_columns(page):
    """Retourne une liste [(x0_debut_colonne, date), ...] triee par x0, une
    entree par jour de la semaine affiche sur cette page du planning (en se
    basant sur les colonnes "Lundi 29/06", "Mardi 30/06", ... et le sous-en-tete
    "Matin" qui marque le debut de chaque colonne jour). Retourne [] si les
    en-tetes attendus ne sont pas trouves (le filtrage par date est alors
    simplement desactive, sans erreur).

    NB: la position verticale (top) de ces en-tetes n'est PAS fixe d'une page
    a l'autre du meme PDF : certains plannings multi-pages redecalent
    legerement tout le contenu sur les pages suivantes (quelques points de
    plus ou de moins). D'anciennes bornes verticales absolues (ex: "60 <= top
    <= 75") marchaient sur la page 1 mais ratissaient a cote sur la page 2+,
    desactivant silencieusement le filtrage par date pour les employes situes
    sur ces pages-la (leurs heures redevenaient alors comptees sur TOUTE la
    semaine du planning au lieu de la seule periode Preparation). On localise
    donc ces en-tetes dynamiquement, relativement les uns aux autres, plutot
    que par une position absolue sur la page."""
    words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
    header_words = [w for w in words if w['top'] < 90]
    if not header_words:
        return []

    header_text = ' '.join(w['text'] for w in sorted(header_words, key=lambda w: (w['top'], w['x0'])))
    year_m = re.search(r'(\d{2})/(\d{2})/(\d{4})', header_text)
    if not year_m:
        return []
    year = int(year_m.group(3))

    # Ligne des jours ("Lundi 29/06", ...) : on la localise via le top des
    # noms de jours eux-memes (une seule occurrence de chaque jour sur la
    # page), sans supposer sa position absolue.
    weekday_tops = sorted({round(w['top'], 1) for w in words if w['text'] in _WEEKDAYS})
    if not weekday_tops:
        return []
    day_top = weekday_tops[0]
    day_row = sorted(
        [w for w in words if abs(w['top'] - day_top) <= 2.5],
        key=lambda w: w['x0'])
    dates = []
    i = 0
    while i < len(day_row):
        w = day_row[i]
        if w['text'] in _WEEKDAYS and i + 1 < len(day_row):
            dm = _DATE_DM_RE.match(day_row[i + 1]['text'])
            if dm:
                try:
                    dates.append(datetime.date(year, int(dm.group(2)), int(dm.group(1))))
                except ValueError:
                    dates.append(None)
                i += 2
                continue
        i += 1

    # Sous-en-tete "Matin" : toujours quelques points sous la ligne des jours
    # (observe ~11pt plus bas), mais ce decalage peut legerement varier d'une
    # page a l'autre -> on prend la ligne de "Matin" la plus proche sous
    # day_top plutot qu'une fenetre absolue fixe.
    matin_candidates = [w for w in words if w['text'] == 'Matin' and w['top'] > day_top]
    if not matin_candidates:
        return []
    matin_top = min(round(w['top'], 1) for w in matin_candidates)
    matin_row = sorted(
        [w for w in matin_candidates if abs(w['top'] - matin_top) <= 2.5],
        key=lambda w: w['x0'])
    starts = [w['x0'] for w in matin_row]

    if not starts or len(starts) != len(dates) or any(d is None for d in dates):
        return []
    return list(zip(starts, dates))


def _date_for_x(x0, day_columns):
    """Renvoie la date de la colonne jour dans laquelle x0 tombe, a partir
    d'une liste [(x0_debut, date), ...] triee par x0 croissant."""
    chosen = None
    for start_x, d in day_columns:
        if x0 >= start_x - 3:
            chosen = d
        else:
            break
    return chosen


def parse_planning_pdf(pdf_path, department='DRIVE', date_start=None, date_end=None):
    """Lit un planning hebdomadaire (format Boulpat/Drive/Bazar) et retourne un
    dict {nom_planning: heures_decimales} ne comptabilisant que les heures
    dont le rayon affecté correspond exactement a `department` (ex: DRIVE).
    Les employes absents toute la semaine (maladie/accident) ou n'ayant
    jamais travaille sur ce rayon obtiennent 0.

    Si date_start/date_end sont fournis (periode reelle du fichier
    Preparation), seules les heures dont la colonne jour du planning tombe
    dans cette periode sont comptabilisees — utile quand le planning couvre
    une semaine calendaire complete (ex: Lundi a Dimanche) mais que le suivi
    de productivite ne porte que sur une partie de cette semaine (ex:
    Mercredi a Dimanche pour la premiere semaine d'un nouveau cycle)."""
    if pdfplumber is None:
        raise RuntimeError("pdfplumber n'est pas installe: impossible de lire le planning PDF.")

    result = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            rows = _cluster_rows(words)
            day_columns = _extract_day_columns(page) if (date_start or date_end) else []

            name_rows_idx = [
                i for i, r in enumerate(rows)
                if r[0]['x0'] < 135 and _NAME_RE.match(r[0]['text']) and len(r[0]['text']) >= 2 and r[0]['top'] > 85
            ]
            for bi, idx in enumerate(name_rows_idx):
                end_idx = name_rows_idx[bi + 1] if bi + 1 < len(name_rows_idx) else len(rows)
                block_rows = rows[idx:end_idx]
                name_tokens = [w['text'] for w in block_rows[0] if w['x0'] < 135]
                name = ' '.join(name_tokens)

                hours_row, tags_row = None, None
                for r in block_rows:
                    toks = [w['text'] for w in r]
                    if toks and all(_HOUR_RE.match(t) for t in toks) and r[0]['x0'] >= 135:
                        hours_row = r
                    elif toks and all(t in _REPOS_TOKENS for t in toks):
                        pass
                    elif toks and all(_TAG_RE.match(t) for t in toks) and r[0]['x0'] >= 135:
                        tags_row = r

                if not hours_row:
                    result[name] = 0.0
                    continue

                hour_tokens = sorted(hours_row, key=lambda w: w['x0'])
                groups = _group_tags(tags_row) if tags_row else []

                dept_hours = 0.0
                for i, htok in enumerate(hour_tokens):
                    tag = groups[i]['text'].strip().upper() if i < len(groups) else None
                    if tag != department.upper():
                        continue
                    if day_columns:
                        d = _date_for_x(htok['x0'], day_columns)
                        if d is not None and date_start is not None and d < date_start:
                            continue
                        if d is not None and date_end is not None and d > date_end:
                            continue
                    dept_hours += _hour_to_decimal(htok['text'])
                result[name] = round(dept_hours, 2)

    return result


def _normalize_name_tokens(name):
    name = re.sub(r'\([^)]*\)', '', name)  # retire le matricule entre parentheses
    name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
    name = re.sub(r'[^A-Za-z\s\-]', ' ', name)
    return [t for t in name.upper().replace('-', ' ').split() if t]


def _name_match_score(name_a, name_b):
    a_tokens = _normalize_name_tokens(name_a)
    b_tokens = _normalize_name_tokens(name_b)
    if not a_tokens or not b_tokens:
        return 0.0
    short, long_ = (a_tokens, b_tokens) if len(a_tokens) <= len(b_tokens) else (b_tokens, a_tokens)
    used, scores = set(), []
    for t in short:
        best, best_i = 0.0, None
        for i, u in enumerate(long_):
            if i in used:
                continue
            s = SequenceMatcher(None, t, u).ratio()
            if s > best:
                best, best_i = s, i
        if best_i is not None:
            used.add(best_i)
        scores.append(best)
    return sum(scores) / len(scores)


def extract_matricule(name):
    m = re.search(r'\(([^)]+)\)', name)
    return m.group(1).strip() if m else None


# Alias manuels pour les employes dont le nom dans le fichier Preparation ne
# partage aucun token avec leur nom reel sur le planning (ex: nom d'emprunt /
# place-holder saisi par erreur), ce que le matching flou (SequenceMatcher)
# ne peut par construction pas relier puisqu'il n'y a aucune similarite de
# caracteres a exploiter. Cle: tuple de tokens normalises (voir
# _normalize_name_tokens) tel que le nom apparait dans la Preparation ->
# tuple de tokens normalises tel que le nom apparait sur le planning.
# Ajouter une entree ici des qu'un nouveau cas de ce type est signale.
NAME_ALIASES = {
    ('COUETTE', 'COUETTE'): ('GONZALES', 'MANRUBIO', 'J'),
}


_NAME_MATCH_THRESHOLD = 0.82


def parse_alias_pairs(raw_pairs):
    """Convertit une liste de paires de noms bruts [(nom_preparation,
    nom_planning), ...] (ex: lues depuis une Data Table n8n persistante) en
    dict {tuple_tokens_preparation: tuple_tokens_planning} au format attendu
    par match_planning_hours. C'est le mecanisme qui permet d'apprendre de
    nouveaux alias SANS toucher au code : un alias confirme par Adrien est
    ajoute comme une ligne de donnee (persistante, relue a chaque
    generation), pas comme une constante codee en dur necessitant un
    redeploiement."""
    out = {}
    for prep_name, plan_name in raw_pairs or []:
        if not prep_name or not plan_name:
            continue
        out[tuple(_normalize_name_tokens(prep_name))] = tuple(_normalize_name_tokens(plan_name))
    return out


def match_planning_hours(planning_hours, employee_names, threshold=_NAME_MATCH_THRESHOLD, extra_aliases=None):
    """Associe chaque employe (nom tel qu'il figure dans la Preparation) aux
    heures Drive du planning, en tolerant fautes de frappe et ordre nom/prenom
    inverse. Ne renvoie une valeur que pour les employes matches avec un score
    suffisant ; les autres sont absents du dict retourne (fallback manuel).

    Verifie d'abord les alias connus (match exact garanti pour les cas ou le
    nom Preparation et le nom planning n'ont rien en commun) avant de
    retomber sur le matching flou pour tous les autres employes. `extra_aliases`
    (deja au format tuple_tokens -> tuple_tokens, voir parse_alias_pairs) est
    fusionne par-dessus les alias codes en dur NAME_ALIASES et les prevaut en
    cas de doublon — c'est la table persistante (Data Table n8n) qui doit
    gagner, puisqu'elle peut etre corrigee sans toucher au code."""
    aliases = dict(NAME_ALIASES)
    if extra_aliases:
        aliases.update(extra_aliases)

    normalized_planning = {}
    for plan_name, hours in planning_hours.items():
        normalized_planning.setdefault(tuple(_normalize_name_tokens(plan_name)), hours)

    matched = {}
    for emp_name in employee_names:
        emp_tokens = tuple(_normalize_name_tokens(emp_name))
        alias_tokens = aliases.get(emp_tokens)
        if alias_tokens is not None and alias_tokens in normalized_planning:
            matched[emp_name] = normalized_planning[alias_tokens]
            continue

        best_score, best_hours = 0.0, None
        for plan_name, hours in planning_hours.items():
            score = _name_match_score(plan_name, emp_name)
            if score > best_score:
                best_score, best_hours = score, hours
        if best_score >= threshold:
            matched[emp_name] = best_hours
    return matched


def build(path, outpath, taux_actuelle=None, taux_precedente=None, planning_path=None, productivite_s1=None,
          name_aliases=None):
    """productivite_s1: dict {matricule: productivite_h_decimale} issu de la
    semaine precedente, utilise pour auto-remplir la colonne H. name_aliases:
    alias supplementaires (deja au format tuple_tokens -> tuple_tokens, voir
    parse_alias_pairs) issus d'une source persistante (Data Table n8n) plutot
    que codes en dur, fusionnes par-dessus NAME_ALIASES. Retourne (semaine,
    nb_employes, productivite_calculee) ou productivite_calculee est un dict
    {matricule: productivite_h_decimale} pour cette semaine, a persister pour
    servir de S-1 la semaine suivante."""
    week, employees = get_week_and_employees(path)
    periode_debut, periode_fin = get_periode(path)

    hours_map = {}
    if planning_path:
        # Ne compter que les heures DRIVE dont la date tombe dans la periode
        # reellement couverte par le fichier Preparation : le planning peut
        # afficher une semaine calendaire complete (Lundi-Dimanche) alors que
        # le suivi de productivite ne porte que sur une partie de celle-ci
        # (ex: premiere semaine d'un cycle qui demarre un mercredi).
        planning_hours = parse_planning_pdf(planning_path, date_start=periode_debut, date_end=periode_fin)
        hours_map = match_planning_hours(planning_hours, [e[0] for e in employees], extra_aliases=name_aliases)

    productivite_s1 = productivite_s1 or {}
    employee_productivity_this_week = {}

    # Pre-clean: round-trip once through openpyxl to strip Numbers-export style
    # cruft that otherwise makes LibreOffice hang when a chart is added later.
    clean_path = path.replace('.xlsx', '_clean.xlsx')
    wb_clean = openpyxl.load_workbook(path, data_only=False)
    wb_clean.save(clean_path)
    wb = openpyxl.load_workbook(clean_path, data_only=False)

    prep_sheet_name = f'Preparation Semaine {week}'
    prod_sheet_name = f'Feuil2 Productivité Semaine {week}'
    assert len(prod_sheet_name) <= 31, prod_sheet_name

    ws_prep = wb['Preparation']
    ws_prep.title = prep_sheet_name

    ws = wb['Feuil2']
    ws.title = prod_sheet_name

    arial = 'Arial'
    n = len(employees)
    first_data_row = 3
    last_data_row = first_data_row + n - 1

    # Title
    ws['A1'] = 'Tableau 2 - Productivité'
    ws['A1'].font = Font(name=arial, bold=True, size=14)

    headers = ['Employé', 'Heure travaillée', "Nbr d'articles préparés",
               'Nbr de commande préparés', 'Productivité /H', 'Productivité minutes',
               'Nombre de commande à l\'heure']
    for i, h in enumerate(headers):
        c = ws.cell(row=2, column=1+i, value=h)
        c.font = Font(name=arial, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    blue_font = Font(name=arial, color='0000FF')

    for i, (name, prep_row, art_row, articles_count, commandes_count) in enumerate(employees):
        r = first_data_row + i
        matricule = extract_matricule(name)
        ws.cell(row=r, column=1, value=name).font = Font(name=arial)

        b_value = None
        cb = ws.cell(row=r, column=2)
        if name in hours_map:
            b_value = hours_map[name]
            cb.value = b_value
            cb.fill = green_fill
            cb.font = Font(name=arial)
        else:
            cb.value = None
            cb.fill = yellow_fill
            cb.font = blue_font
        cb.number_format = '0.00'

        # C/D/E/F/G : toujours des formules Excel live (references a la feuille
        # Preparation et a la colonne B), meme quand B est deja auto-rempli
        # depuis le planning. Ainsi, si l'utilisateur modifie une heure (ou
        # tout autre nombre) dans Excel, la productivite et le graphique se
        # recalculent automatiquement. Le fichier livre est tout de meme
        # recalcule cote serveur (LibreOffice) avant envoi, donc les valeurs
        # mises en cache sont correctes des l'ouverture, meme dans un lecteur
        # qui n'executerait pas le recalcul automatiquement.
        e_value = (articles_count / b_value) if (articles_count is not None and b_value) else None
        cc = ws.cell(row=r, column=3, value=f"='{prep_sheet_name}'!C{art_row}")
        cd = ws.cell(row=r, column=4, value=f"='{prep_sheet_name}'!C{prep_row}")
        ce = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},\"\")")
        cf = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/60,\"\")")
        cg = ws.cell(row=r, column=7, value=f"=IFERROR(D{r}/B{r},\"\")")
        cc.font = Font(name=arial, color='008000')
        cd.font = Font(name=arial, color='008000')
        ce.font = Font(name=arial)
        ce.number_format = '0.00'
        cf.font = Font(name=arial)
        cf.number_format = '0.00'
        cg.font = Font(name=arial)
        cg.number_format = '0.00'

        h_value = None
        ch = ws.cell(row=r, column=8)
        if matricule and matricule in productivite_s1:
            h_value = productivite_s1[matricule]
            ch.value = h_value
            ch.fill = green_fill
            ch.font = Font(name=arial)
        else:
            ch.value = None
            ch.fill = yellow_fill
            ch.font = blue_font
        ch.number_format = '0.00'

        # % Evolution vs S-1 : difference entre Productivite/H (E, semaine
        # actuelle) et Productivite/H S-1 (H), exprimee en % de la valeur
        # ACTUELLE (E) — donc (E-H)/E*100, pas /H. Toujours une formule live
        # (meme raison que C-G ci-dessus : rester dynamique si l'utilisateur
        # modifie les heures ou la productivite S-1 dans Excel).
        ci = ws.cell(row=r, column=9,
                      value=(f"=IFERROR(IF(OR(E{r}=\"\",H{r}=\"\"),\"\",IF(E{r}>=H{r},\"▲ \",\"▼ \")"
                             f"&ROUND(ABS(E{r}-H{r})/E{r}*100,1)&\"%\"),\"\")"))
        ci.font = Font(name=arial, bold=True)
        ci.alignment = Alignment(horizontal='center')

        # Productivite calculee cette semaine (pour servir de S-1 la semaine prochaine)
        # -- calcul Python interne uniquement, n'affecte pas les cellules ecrites.
        if matricule and b_value and articles_count is not None:
            entry = {
                'nom': name,
                'productivite_h': round(articles_count / b_value, 2),
            }
            if e_value is not None and h_value is not None and e_value != 0:
                entry['evolution_pct'] = round(abs(e_value - h_value) / e_value * 100, 1)
                entry['evolution_up'] = e_value >= h_value
            employee_productivity_this_week[matricule] = entry

    headers2 = ['Productivité /H S-1', '% Évolution vs S-1']
    for i, h in enumerate(headers2):
        c = ws.cell(row=2, column=8+i, value=h)
        c.font = Font(name=arial, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    ws.conditional_formatting.add(
        f'I{first_data_row}:I{last_data_row}',
        FormulaRule(formula=[f'AND($E{first_data_row}<>\"\",$H{first_data_row}<>\"\",$E{first_data_row}>=$H{first_data_row})'],
                    font=Font(name=arial, color='008000', bold=True)))
    ws.conditional_formatting.add(
        f'I{first_data_row}:I{last_data_row}',
        FormulaRule(formula=[f'AND($E{first_data_row}<>\"\",$H{first_data_row}<>\"\",$E{first_data_row}<$H{first_data_row})'],
                    font=Font(name=arial, color='FF0000', bold=True)))

    widths = {'A': 32, 'B': 16, 'C': 20, 'D': 20, 'E': 16, 'F': 16, 'G': 20, 'H': 18, 'I': 18}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    k_headers = ['Rang', 'Employé (classé)', 'Productivité /H (classée)']
    for i, h in enumerate(k_headers):
        c = ws.cell(row=2, column=11+i, value=h)
        c.font = Font(name=arial, bold=True)
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    e_range = f"$E${first_data_row}:$E${last_data_row}"
    a_range = f"$A${first_data_row}:$A${last_data_row}"

    for i in range(n):
        r = first_data_row + i
        ck = ws.cell(row=r, column=11, value=f"=ROW()-{first_data_row-1}")
        ck.font = Font(name=arial)
        cm = ws.cell(row=r, column=13, value=f"=IFERROR(LARGE({e_range},K{r}),\"\")")
        cm.font = Font(name=arial)
        cm.number_format = '0.00'
        cl = ws.cell(row=r, column=12,
                      value=f"=IF(M{r}=\"\",\"\",IFERROR(INDEX({a_range},MATCH(M{r},{e_range},0)),\"\"))")
        cl.font = Font(name=arial)

    widths2 = {'K': 8, 'L': 32, 'M': 22}
    for col, w in widths2.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('O1:P1')
    ct = ws['O1']
    ct.value = 'Indicateur - Taux de rupture (avant substitution)'
    ct.font = Font(name=arial, bold=True)
    ct.alignment = Alignment(horizontal='center', wrap_text=True)

    ws['O2'] = 'Semaine précédente (S-1)'
    ws['O2'].font = Font(name=arial)
    p2 = ws['P2']
    p2.fill = yellow_fill
    p2.font = blue_font
    p2.number_format = '0.0%'

    ws['O3'] = 'Semaine actuelle'
    ws['O3'].font = Font(name=arial)
    p3 = ws['P3']
    p3.fill = yellow_fill
    p3.font = blue_font
    p3.number_format = '0.0%'
    if taux_actuelle is not None:
        p3.value = taux_actuelle
    if taux_precedente is not None:
        p2.value = taux_precedente

    ws['O4'] = 'Évolution'
    ws['O4'].font = Font(name=arial, bold=True)
    p4 = ws['P4']
    p4.value = ('=IFERROR(IF(OR(P2="",P3=""),"",ROUND(P2*100,1)&"%"&"  "'
                '&IF(P3<=P2,"▼","▲")&"  "&ROUND(P3*100,1)&"%"),"")')
    p4.font = Font(name=arial, bold=True)
    p4.alignment = Alignment(horizontal='center')

    ws.conditional_formatting.add(
        'P4', FormulaRule(formula=['AND(P2<>"",P3<>"",P3<=P2)'],
                           font=Font(name=arial, color='008000', bold=True, italic=True)))
    ws.conditional_formatting.add(
        'P4', FormulaRule(formula=['AND(P2<>"",P3<>"",P3>P2)'],
                           font=Font(name=arial, color='FF0000', bold=True, italic=True)))

    widths3 = {'O': 24, 'P': 20}
    for col, w in widths3.items():
        ws.column_dimensions[col].width = w

    wb.calculation.fullCalcOnLoad = True
    wb.save(outpath)
    wb = openpyxl.load_workbook(outpath)
    ws = wb[prod_sheet_name]

    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Productivité à l\'heure par employé (du plus grand au plus petit)'
    chart.y_axis.title = 'Productivité /H'
    chart.x_axis.title = 'Employé'
    chart.style = 10
    chart.width = max(24, n * 2.2)
    chart.height = 12

    data = Reference(ws, min_col=13, min_row=2, max_row=last_data_row)
    cats = Reference(ws, min_col=12, min_row=first_data_row, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    cat_str_ref = StrRef(f=str(cats))
    chart.series[0].cat = AxDataSource(strRef=cat_str_ref)

    # Une couleur distincte par employe (barre), plutot qu'une seule couleur
    # pour toute la serie.
    palette = [
        '4472C4', 'ED7D31', 'A5A5A5', 'FFC000', '5B9BD5', '70AD47',
        '264478', '9E480E', '636363', '997300', '255E91', '43682B',
        'C00000', '7030A0', '00B0F0', 'FF66CC', '00B050', 'BF8F00',
        '203864', '833C00',
    ]
    chart.series[0].graphicalProperties.varyColors = True
    data_points = []
    for idx in range(n):
        color = palette[idx % len(palette)]
        dp = DataPoint(idx=idx)
        dp.graphicalProperties = GraphicalProperties(solidFill=color)
        data_points.append(dp)
    chart.series[0].data_points = data_points

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.showLegendKey = False
    chart.dLbls.showCatName = False
    chart.dLbls.showSerName = False
    chart.legend = None

    body_pr = RichTextProperties(rot=-2700000, vert='horz')
    cp = CharacterProperties(sz=900)
    pp = ParagraphProperties(defRPr=cp)
    rich = RichText(bodyPr=body_pr, p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chart.x_axis.txPr = rich
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    chart_anchor_row = last_data_row + 4
    ws.add_chart(chart, f'A{chart_anchor_row}')

    wb.calculation.fullCalcOnLoad = True
    wb.save(outpath)
    return week, n, employee_productivity_this_week


# ---------------------------------------------------------------------------
# PDF "Podium" - affichage visuel hebdomadaire (charte Intermarché) a afficher
# en salle de pause. Design valide avec Adrien le 02/08/2026.
# ---------------------------------------------------------------------------

_PDM_RED = None
_PDM_BLACK = None


def _pdm_colors():
    global _PDM_RED, _PDM_BLACK
    if _PDM_RED is None:
        _PDM_RED = _rl_colors.HexColor('#E2001A')
        _PDM_BLACK = _rl_colors.HexColor('#1A1A1A')
    return _PDM_RED, _PDM_BLACK


def _pdm_rounded_card(c, x, y, w, h, fill, stroke=None, radius=10, stroke_width=1.2):
    c.saveState()
    c.setFillColor(fill)
    if stroke:
        c.setStrokeColor(stroke)
        c.setLineWidth(stroke_width)
        c.roundRect(x, y, w, h, radius, fill=1, stroke=1)
    else:
        c.roundRect(x, y, w, h, radius, fill=1, stroke=0)
    c.restoreState()


def _pdm_center_text(c, text, cx, y, font, size, color):
    c.setFont(font, size)
    c.setFillColor(color)
    c.drawCentredString(cx, y, text)


def _pdm_evolution_pill(c, cx, y, up, pct, WHITE, GREEN, RED_NEG, GREY):
    if pct is None:
        label, color = "NOUVEAU", GREY
    else:
        label = f"{'▲' if up else '▼'} {pct:.1f}%"
        color = GREEN if up else RED_NEG
    c.setFont("Helvetica-Bold", 10)
    tw = _rl_stringWidth(label, "Helvetica-Bold", 10)
    pad = 10
    pw, ph = tw + 2 * pad, 18
    _pdm_rounded_card(c, cx - pw / 2, y, pw, ph, color, radius=9)
    c.setFillColor(WHITE)
    c.drawCentredString(cx, y + 5, label)


def _pdm_medal_circle(c, cx, cy, r, color, text, WHITE):
    c.setFillColor(color)
    c.circle(cx, cy, r, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", r * 0.85)
    c.drawCentredString(cx, cy - r * 0.32, text)


def _pdm_logo_mark(c, x, y, size, RED, WHITE):
    """Badge enseigne : carre rouge arrondi + panier stylise (icone generique)."""
    _pdm_rounded_card(c, x, y, size, size, RED, radius=size * 0.22)
    c.saveState()
    cx, cy = x + size / 2, y + size / 2 - size * 0.04
    c.setStrokeColor(WHITE)
    c.setLineWidth(size * 0.055)
    c.arc(cx - size * 0.16, cy + size * 0.02, cx + size * 0.16, cy + size * 0.32, 0, 180)
    basket_w, basket_h = size * 0.46, size * 0.24
    bx, by = cx - basket_w / 2, cy - basket_h / 2
    p = c.beginPath()
    p.moveTo(bx + basket_w * 0.08, by + basket_h)
    p.lineTo(bx, by)
    p.lineTo(bx + basket_w, by)
    p.lineTo(bx + basket_w * 0.92, by + basket_h)
    p.close()
    c.setFillColor(WHITE)
    c.drawPath(p, fill=1, stroke=0)
    c.circle(bx + basket_w * 0.2, by - size * 0.05, size * 0.035, fill=1, stroke=0)
    c.circle(bx + basket_w * 0.8, by - size * 0.05, size * 0.035, fill=1, stroke=0)
    c.restoreState()


def _pdm_clean_name(nom):
    base = re.sub(r'\([^)]*\)', '', nom).strip()
    return base.title()


def build_ranking(employee_productivity, excluded_matricules=None):
    """Construit la liste classee (desc. par productivite_h) a partir du dict
    retourne par build(), en excluant les responsables."""
    excluded = excluded_matricules if excluded_matricules is not None else EXCLUDED_FROM_RANKING
    rows = []
    for matricule, data in (employee_productivity or {}).items():
        if matricule in excluded:
            continue
        if data.get('productivite_h') is None:
            continue
        if data.get('productivite_h') <= 0:
            # Un 0h signale presque toujours une anomalie de matricule (ex: la
            # meme personne saisie sous deux matricules differents une
            # semaine donnee) plutot qu'un veritable collaborateur a classer.
            continue
        rows.append({
            'matricule': matricule,
            'name': _pdm_clean_name(data.get('nom', matricule)),
            'prod': data['productivite_h'],
            'evo': data.get('evolution_pct'),
            'up': data.get('evolution_up'),
        })
    rows.sort(key=lambda r: r['prod'], reverse=True)
    return rows


def generate_podium_pdf(pdf_path, week, employee_productivity, taux_actuelle=None,
                         taux_precedente=None, enseigne="INTERMARCHÉ", magasin="MONTESCOT",
                         excluded_matricules=None):
    """Genere le PDF "podium" hebdomadaire (charte Intermarche) a partir des
    resultats de build(). Retourne le nombre de collaborateurs classes."""
    if _rl_canvas is None:
        raise RuntimeError("reportlab n'est pas installe: impossible de generer le PDF podium.")

    RED, BLACK = _pdm_colors()
    WHITE = _rl_colors.white
    GOLD = _rl_colors.HexColor('#D9A400')
    SILVER = _rl_colors.HexColor('#9AA0A6')
    BRONZE = _rl_colors.HexColor('#B5651D')
    GREEN = _rl_colors.HexColor('#1E7B34')
    RED_NEG = _rl_colors.HexColor('#C62828')
    GREY_BG = _rl_colors.HexColor('#F3F3F3')
    GREY_TXT = _rl_colors.HexColor('#5B5B5B')
    GREY_PILL = _rl_colors.HexColor('#9E9E9E')

    ranking = build_ranking(employee_productivity, excluded_matricules)
    W, H = A4
    c = _rl_canvas.Canvas(pdf_path, pagesize=A4)

    # ---- Header
    header_h = 108
    c.setFillColor(BLACK)
    c.rect(0, H - header_h, W, header_h, fill=1, stroke=0)
    c.setFillColor(RED)
    c.rect(0, H - header_h - 6, W, 6, fill=1, stroke=0)

    logo_size = 52
    logo_x, logo_y = 36, H - 34 - logo_size
    _pdm_logo_mark(c, logo_x, logo_y, logo_size, RED, WHITE)

    text_x = logo_x + logo_size + 14
    c.setFont("Helvetica-Bold", 25)
    c.setFillColor(WHITE)
    c.drawString(text_x, H - 46, enseigne)
    c.setFillColor(RED)
    tw = _rl_stringWidth(enseigne + " ", "Helvetica-Bold", 25)
    c.drawString(text_x + tw, H - 46, magasin)

    c.setFont("Helvetica-Bold", 11)
    band_label = "PERFORMANCE DRIVE"
    btw = _rl_stringWidth(band_label, "Helvetica-Bold", 11)
    bpad = 8
    _pdm_rounded_card(c, text_x, H - 74, btw + 2 * bpad, 17, RED, radius=8)
    c.setFillColor(WHITE)
    c.drawString(text_x + bpad, H - 70, band_label)
    c.setFont("Helvetica", 10)
    c.setFillColor(_rl_colors.HexColor('#CCCCCC'))
    c.drawString(text_x + btw + 2 * bpad + 10, H - 70, "Classement hebdomadaire des collaborateurs")

    pill_label = f"SEMAINE {week}"
    c.setFont("Helvetica-Bold", 13)
    ptw = _rl_stringWidth(pill_label, "Helvetica-Bold", 13)
    pw, ph = ptw + 28, 30
    # Ancre en haut a droite (aligne sur le titre de l'enseigne) plutot que
    # centre sur toute la hauteur de l'entete : evite qu'il ne vienne
    # frotter contre le bandeau "PERFORMANCE DRIVE" en dessous, quelle que
    # soit la longueur du nom d'enseigne/magasin.
    px, py = W - 36 - pw, H - 34 - ph
    _pdm_rounded_card(c, px, py, pw, ph, RED, radius=15)
    c.setFillColor(WHITE)
    c.drawCentredString(px + pw / 2, py + ph / 2 - 5, pill_label)

    # ---- Podium top 3
    podium_top = H - header_h - 18
    card_w, gap = 152, 14
    total_w = card_w * 3 + gap * 2
    start_x = (W - total_w) / 2
    order = [1, 0, 2]
    medal_colors = {0: GOLD, 1: SILVER, 2: BRONZE}
    card_h_map = {0: 172, 1: 148, 2: 148}
    bar_h_map = {0: 150, 1: 108, 2: 78}
    baseline = podium_top - 200

    top3 = ranking[:3]
    for slot, idx in enumerate(order):
        if idx >= len(top3):
            continue
        x = start_x + slot * (card_w + gap)
        p = top3[idx]
        card_h = card_h_map[idx]
        card_y = podium_top - card_h
        color = medal_colors[idx]

        _pdm_rounded_card(c, x, card_y, card_w, card_h, WHITE, stroke=color, stroke_width=2.2, radius=12)
        cx = x + card_w / 2
        _pdm_medal_circle(c, cx, card_y + card_h - 8, 22, color, str(idx + 1), WHITE)

        c.setFont("Helvetica-Bold", 11)
        name = p["name"]
        if _rl_stringWidth(name, "Helvetica-Bold", 11) > card_w - 16:
            parts = name.split(" ", 1)
            _pdm_center_text(c, parts[0], cx, card_y + card_h - 46, "Helvetica-Bold", 11, BLACK)
            _pdm_center_text(c, parts[1] if len(parts) > 1 else "", cx, card_y + card_h - 59, "Helvetica-Bold", 11, BLACK)
            name_bottom = card_y + card_h - 59
        else:
            _pdm_center_text(c, name, cx, card_y + card_h - 46, "Helvetica-Bold", 11, BLACK)
            name_bottom = card_y + card_h - 46

        _pdm_center_text(c, f"{p['prod']:.1f}", cx, name_bottom - 24, "Helvetica-Bold", 20, BLACK)
        _pdm_center_text(c, "articles / heure", cx, name_bottom - 36, "Helvetica", 7.5, GREY_TXT)
        _pdm_evolution_pill(c, cx, card_y + 10, p["up"], p["evo"], WHITE, GREEN, RED_NEG, GREY_PILL)

        bar_h = bar_h_map[idx]
        bar_y = baseline - bar_h
        c.setFillColor(color)
        c.rect(x, bar_y, card_w, bar_h, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 34)
        c.setFillColor(WHITE)
        c.drawCentredString(cx, bar_y + bar_h / 2 - 12, str(idx + 1))

    list_top = baseline - max(bar_h_map.values()) - 22

    # ---- Blocs du bas (taux de rupture + ruban meilleure progression),
    # ancres depuis le BAS de la page plutot que depuis le contenu du
    # dessus : le nombre de lignes du classement et la presence ou non
    # d'un ruban varient d'une semaine a l'autre, et un positionnement
    # "en cascade" depuis le haut laissait soit un grand vide (peu de
    # lignes / pas de ruban -> semaine 27), soit un chevauchement avec le
    # pied de page (beaucoup de lignes -> semaine 28). En ancrant ces deux
    # blocs depuis le bas, ils restent toujours a une position fixe et
    # lisible, quel que soit le contenu du classement.
    footer_h = 34
    bottom_margin = 20
    taux_h = 108
    ribbon_h = 46

    positive = [p for p in ranking if p["up"]]
    has_ribbon = bool(positive)

    box_y = footer_h + bottom_margin
    if has_ribbon:
        ribbon_y = box_y
        box_y = ribbon_y + ribbon_h + 16
    box_top = box_y + taux_h

    # ---- Classement 4e et suivants : la table occupe tout l'espace
    # disponible entre le podium et le bloc du bas, avec une hauteur de
    # ligne qui s'adapte (dans des bornes raisonnables) pour ne jamais se
    # chevaucher avec le bas de page ni laisser un vide disproportionne.
    header_row_h = 22
    list_rows = ranking[3:]
    available = list_top - 24 - box_top
    if list_rows:
        # Le plancher precedent (22) etait plus grand que "available" des
        # qu'il y avait beaucoup de collaborateurs classes (ex: semaine 28,
        # 8 collaborateurs -> 5 lignes hors podium) : la table debordait
        # alors mecaniquement sur le bloc "Taux de rupture" en dessous,
        # quel que soit son contenu. Le plancher doit rester en-dessous de
        # ce que "available" peut effectivement fournir ; 12pt reste
        # lisible pour une table dense, et le texte est recentre/reduit
        # dynamiquement ci-dessous quand row_h retrecit.
        row_h = max(12, min(34, (available - header_row_h) / len(list_rows)))
    else:
        row_h = 24
    # Taille de police et decalage vertical du texte adaptes a row_h, pour
    # que chaque ligne reste lisible et centree meme quand la table est
    # tres dense (beaucoup de collaborateurs classes une semaine donnee).
    if row_h >= 20:
        row_font = 10
    elif row_h >= 16:
        row_font = 9
    else:
        row_font = 7.5
    text_y_offset = max(3, min(8, row_h / 2 + row_font / 2 - 2))
    y = list_top
    c.setFillColor(RED)
    c.rect(36, y - header_row_h, W - 72, header_row_h, fill=1, stroke=0)
    c.setFillColor(WHITE)
    c.setFont("Helvetica-Bold", 9.5)
    c.drawString(50, y - header_row_h + 7, "RANG")
    c.drawString(100, y - header_row_h + 7, "COLLABORATEUR")
    c.drawString(360, y - header_row_h + 7, "PRODUCTIVITÉ /H")
    c.drawString(480, y - header_row_h + 7, "ÉVOLUTION S-1")
    y -= header_row_h

    for i, p in enumerate(list_rows):
        rank = i + 4
        row_y = y - row_h
        if i % 2 == 0:
            c.setFillColor(GREY_BG)
            c.rect(36, row_y, W - 72, row_h, fill=1, stroke=0)
        c.setFillColor(BLACK)
        c.setFont("Helvetica-Bold", row_font)
        c.drawString(50, row_y + text_y_offset, f"{rank}e")
        c.setFont("Helvetica", row_font)
        c.drawString(100, row_y + text_y_offset, p["name"])
        c.setFont("Helvetica-Bold", row_font)
        c.drawString(360, row_y + text_y_offset, f"{p['prod']:.1f} art./h")
        if p["evo"] is None:
            c.setFillColor(GREY_PILL)
            c.drawString(480, row_y + text_y_offset, "NOUVEAU")
        else:
            arrow = "▲" if p["up"] else "▼"
            c.setFillColor(GREEN if p["up"] else RED_NEG)
            c.drawString(480, row_y + text_y_offset, f"{arrow} {p['evo']:.1f}%")
        y = row_y

    # Garde-fou structurel : la table de classement ne doit JAMAIS empieter
    # sur le bloc "Taux de rupture" dessine juste apres. row_h est deja
    # calcule pour que ça tienne (voir plus haut), mais si un cas extreme
    # (ex: un tres grand nombre de collaborateurs classes une semaine
    # donnee) fait quand meme deborder malgre le plancher de 12pt, on
    # prefere echouer bruyamment ici plutot que livrer silencieusement un
    # PDF avec un chevauchement visuel — c'est exactement le type de bug
    # remonte sur la semaine 28 avant ce correctif.
    if list_rows and y < box_top - 0.5:
        raise RuntimeError(
            f"Mise en page podium semaine {week} : le classement (jusqu'a y={y:.1f}) "
            f"chevaucherait le bloc Taux de rupture (qui commence a y={box_top:.1f}). "
            f"Trop de collaborateurs classes ({len(list_rows)} lignes hors podium) pour "
            f"tenir dans l'espace disponible meme a hauteur de ligne minimale. "
            f"Generation du PDF annulee — a corriger (ex: pagination, ou reduire le nombre "
            f"de lignes affichees) plutot que livrer un visuel casse."
        )

    # ---- Taux de rupture (position deja fixee plus haut, ancree sur le
    # bas de page : voir box_y/ribbon_y calcules avant la table)
    _pdm_rounded_card(c, 36, box_y, W - 72, taux_h, BLACK, radius=12)
    c.setFillColor(RED)
    c.setFont("Helvetica-Bold", 13)
    c.drawCentredString(W / 2, box_y + taux_h - 24, "TAUX DE RUPTURE")

    half_w = (W - 72) / 2
    tac = taux_actuelle if taux_actuelle is not None else 0.0
    _pdm_center_text(c, "SEMAINE PRÉCÉDENTE (S-1)", 36 + half_w / 2, box_y + taux_h - 50, "Helvetica", 9, _rl_colors.HexColor('#BBBBBB'))
    if taux_precedente is None:
        _pdm_center_text(c, "N/D", 36 + half_w / 2, box_y + 22, "Helvetica-Bold", 26, _rl_colors.HexColor('#777777'))
    else:
        _pdm_center_text(c, f"{taux_precedente*100:.2f}%", 36 + half_w / 2, box_y + 22, "Helvetica-Bold", 26, WHITE)
    _pdm_center_text(c, "CETTE SEMAINE", 36 + half_w + half_w / 2, box_y + taux_h - 50, "Helvetica", 9, _rl_colors.HexColor('#BBBBBB'))

    ts1 = taux_precedente
    delta = (tac - ts1) if ts1 is not None else None
    worse = delta is not None and delta > 0
    delta_color = (RED_NEG if worse else GREEN) if delta is not None else WHITE
    _pdm_center_text(c, f"{tac*100:.2f}%", 36 + half_w + half_w / 2, box_y + 22, "Helvetica-Bold", 26, delta_color)

    c.setStrokeColor(_rl_colors.HexColor('#444444'))
    c.setLineWidth(1)
    c.line(36 + half_w, box_y + 14, 36 + half_w, box_y + taux_h - 40)

    if delta is not None:
        delta_label = f"{'+' if worse else ''}{delta*100:.2f} pt {'▲' if worse else '▼'}"
        c.setFont("Helvetica-Bold", 10)
        dtw = _rl_stringWidth(delta_label, "Helvetica-Bold", 10)
        dpw = dtw + 20
        _pdm_rounded_card(c, W / 2 - dpw / 2, box_y + taux_h - 46, dpw, 17, delta_color, radius=8)
        c.setFillColor(WHITE)
        c.drawCentredString(W / 2, box_y + taux_h - 46 + 4.5, delta_label)

    # ---- Meilleure progression (position deja fixee plus haut : ribbon_y
    # n'existe que si has_ribbon est vrai)
    if has_ribbon:
        best = max(positive, key=lambda p: p["evo"])
        _pdm_rounded_card(c, 36, ribbon_y, W - 72, ribbon_h, RED, radius=10)
        c.setFillColor(WHITE)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(W / 2, ribbon_y + ribbon_h / 2 + 4,
                             f"★ MEILLEURE PROGRESSION DE LA SEMAINE : {best['name']} (▲ {best['evo']:.1f}%)")

    # ---- Footer
    footer_h = 34
    c.setFillColor(BLACK)
    c.rect(0, 0, W, footer_h, fill=1, stroke=0)
    c.setFillColor(RED)
    c.rect(0, footer_h - 3, W, 3, fill=1, stroke=0)
    today = datetime.date.today().strftime("%d/%m/%Y")
    c.setFont("Helvetica", 9)
    c.setFillColor(WHITE)
    c.drawCentredString(W / 2, footer_h / 2 - 3,
                         f"Semaine {week}  ·  Généré le {today}  ·  Merci pour votre engagement au quotidien !")

    c.save()
    return len(ranking)


# ---------------------------------------------------------------------------
# Verification post-generation (agent de controle qualite) - recontrole,
# de maniere independante, ce que build() vient de produire, pour detecter
# avant l'envoi les classes de bugs deja rencontrees en production :
# heures incoherentes avec le planning, employe non retrouve automatiquement
# (alias manquant), formules figees au lieu de formules dynamiques, fichier
# corrompu au recalcul. Ne leve pas d'exception : retourne un rapport que
# l'appelant (endpoint HTTP / n8n) utilise pour decider quoi faire (bloquer,
# alerter en parallele, envoyer quand meme...).
# ---------------------------------------------------------------------------

_XLSX_ERROR_RE = re.compile(r'^#(DIV/0!|N/A|NAME\?|NULL!|NUM!|REF!|VALUE!)$')


def _soffice_binary():
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def _recalc_with_libreoffice(xlsx_path, timeout=45):
    """Tente un recalcul via LibreOffice headless (meme principe que le
    script recalc.py utilise pendant le developpement). Retourne (ok, msg) :
    ok=True si le recalcul a reussi sans erreur de formule (le fichier est
    alors reecrit avec les valeurs mises en cache, utile pour les
    visionneuses qui n'executent pas de recalcul automatique comme
    l'apercu Google Drive) ; ok=None si LibreOffice n'est pas installe sur
    ce serveur (verification sautee — PAS un echec en soi, puisque
    fullCalcOnLoad force Excel a recalculer a l'ouverture de toute facon) ;
    ok=False si des erreurs de formule ont ete trouvees ou si la conversion
    a echoue."""
    soffice = _soffice_binary()
    if not soffice:
        return None, ("LibreOffice non installe sur ce serveur : recalcul non verifie "
                       "(Excel recalculera correctement a l'ouverture grace a fullCalcOnLoad, "
                       "mais un apercu Drive/Docs peut rester vide tant que le fichier n'a pas "
                       "ete ouvert au moins une fois dans Excel).")
    out_dir = tempfile.mkdtemp(prefix="recalc_")
    try:
        proc = subprocess.run(
            [soffice, "--headless", "--norestore", "--convert-to",
             "xlsx:Calc MS Excel 2007 XML", "--outdir", out_dir, xlsx_path],
            capture_output=True, timeout=timeout,
        )
        converted = os.path.join(out_dir, os.path.basename(xlsx_path))
        if proc.returncode != 0 or not os.path.exists(converted):
            return False, f"LibreOffice a echoue (code {proc.returncode}): {proc.stderr.decode(errors='ignore')[:500]}"
        wb = openpyxl.load_workbook(converted, data_only=True)
        bad = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and _XLSX_ERROR_RE.match(cell.value):
                        bad.append(f"{ws.title}!{cell.coordinate}={cell.value}")
        if bad:
            return False, f"{len(bad)} erreur(s) de formule apres recalcul: " + ", ".join(bad[:10])
        shutil.copyfile(converted, xlsx_path)
        return True, "Recalcul LibreOffice OK, 0 erreur de formule."
    except subprocess.TimeoutExpired:
        return False, f"Recalcul LibreOffice : timeout ({timeout}s)."
    except Exception as e:
        return False, f"Recalcul LibreOffice : erreur inattendue ({e})."
    finally:
        shutil.rmtree(out_dir, ignore_errors=True)


def verify_output(prep_path, xlsx_path, planning_path=None, name_aliases=None, recalc=True):
    """Controle qualite du fichier "Performance Drive" genere, avant envoi.

    Recalcule independamment les heures Drive depuis le planning et les
    compare a ce qui a ete ecrit dans le fichier (au lieu de faire confiance
    aveuglement au resultat de build()), verifie que les colonnes calculees
    restent des formules dynamiques (pas des valeurs figees), signale tout
    employe que le matching automatique n'a pas retrouve dans le planning
    (avec le nom le plus proche trouve, pour faciliter l'ajout d'un alias),
    et tente un recalcul LibreOffice pour detecter une erreur de formule.

    Retourne {"ok": bool, "errors": [...], "warnings": [...], "week": int}.
    - errors : problemes qui remettent en cause l'exactitude des donnees ou
      cassent le fichier -> ne doit jamais partir tel quel (bloquant).
    - warnings : points a verifier par un humain (ex: employe non retrouve
      automatiquement) mais qui n'invalident pas le fichier -> peut partir,
      avec une notification en parallele plutot qu'un blocage."""
    errors, warnings = [], []

    week, employees = get_week_and_employees(prep_path)
    periode_debut, periode_fin = get_periode(prep_path)

    hours_map, planning_hours = {}, {}
    if planning_path:
        planning_hours = parse_planning_pdf(planning_path, date_start=periode_debut, date_end=periode_fin)
        hours_map = match_planning_hours(planning_hours, [e[0] for e in employees], extra_aliases=name_aliases)

    prod_sheet_name = f'Feuil2 Productivité Semaine {week}'

    wb_f = openpyxl.load_workbook(xlsx_path, data_only=False)
    if prod_sheet_name not in wb_f.sheetnames:
        errors.append(f"Feuille '{prod_sheet_name}' absente du fichier genere (attendu pour la semaine {week}).")
        return {"ok": False, "errors": errors, "warnings": warnings, "week": week}
    ws_f = wb_f[prod_sheet_name]

    wb_v = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_v = wb_v[prod_sheet_name]

    for i, (name, prep_row, art_row, articles_count, commandes_count) in enumerate(employees):
        r = 3 + i
        cell_name = ws_v.cell(row=r, column=1).value
        if cell_name != name:
            errors.append(f"Ligne {r}: nom attendu '{name}', trouve '{cell_name}' (desalignement des lignes).")
            continue

        if planning_path:
            b_actual = ws_v.cell(row=r, column=2).value
            b_expected = hours_map.get(name)
            if b_expected is not None:
                if b_actual is None or abs(float(b_actual) - float(b_expected)) > 0.01:
                    errors.append(
                        f"{name}: heures ecrites dans le fichier ({b_actual}) different des heures "
                        f"recalculees independamment depuis le planning ({b_expected}) sur la periode "
                        f"{periode_debut}..{periode_fin}."
                    )
            else:
                # Pas de correspondance automatique : comportement normal
                # (cellule laissee vide pour saisie manuelle), sauf si un nom
                # de planning suffisamment proche existe malgre tout — signe
                # possible d'un alias manquant (cas "Couette Couette /
                # Gonzales-Manrubio J").
                best_score, best_plan_name = 0.0, None
                for plan_name in planning_hours:
                    s = _name_match_score(plan_name, name)
                    if s > best_score:
                        best_score, best_plan_name = s, plan_name
                if best_plan_name and 0.25 <= best_score < _NAME_MATCH_THRESHOLD:
                    warnings.append(
                        f"{name}: aucune heure trouvee automatiquement dans le planning. Nom le plus "
                        f"proche trouve : '{best_plan_name}' ({planning_hours[best_plan_name]}h, score de "
                        f"similarite {best_score:.2f} — trop different pour un rapprochement automatique "
                        f"fiable). Si c'est la meme personne, confirmer pour ajouter l'alias."
                    )
                elif b_actual is None:
                    warnings.append(f"{name}: aucune heure trouvee dans le planning, saisie manuelle necessaire.")

        for col, label in ((3, 'C'), (4, 'D'), (5, 'E'), (6, 'F'), (7, 'G'), (9, 'I')):
            v = ws_f.cell(row=r, column=col).value
            if not (isinstance(v, str) and v.startswith('=')):
                errors.append(
                    f"{name}, colonne {label}{r}: valeur figee au lieu d'une formule ({v!r}) — le fichier "
                    f"ne se recalculera plus si l'utilisateur modifie ses heures dans Excel."
                )

    if not ws_f._charts:
        errors.append("Aucun graphique trouve sur la feuille de productivite.")

    if recalc:
        recalc_ok, recalc_msg = _recalc_with_libreoffice(xlsx_path)
        if recalc_ok is False:
            errors.append(recalc_msg)
        elif recalc_ok is None:
            warnings.append(recalc_msg)

    return {"ok": len(errors) == 0, "errors": errors, "warnings": warnings, "week": week}


def verify_podium_pdf(pdf_path, expected_min_collaborateurs=1):
    """Controle qualite du PDF podium : le generateur (generate_podium_pdf)
    leve deja une RuntimeError si le classement chevaucherait le bloc Taux
    de rupture, donc un PDF qui existe a deja passe ce garde-fou structurel.
    Cette fonction complete par un controle generique, independant de la
    logique de dessin : aucun texte rendu ne doit se chevaucher visuellement
    avec un autre, quelle que soit la section concernee — de quoi attraper
    une future regression de mise en page ailleurs sur la page, pas
    seulement dans le classement."""
    errors, warnings = [], []
    if pdfplumber is None:
        warnings.append("pdfplumber non installe : verification visuelle du PDF sautee.")
        return {"ok": True, "errors": errors, "warnings": warnings}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                errors.append("Le PDF genere ne contient aucune page.")
                return {"ok": False, "errors": errors, "warnings": warnings}
            page = pdf.pages[0]
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            if len(words) < 10:
                errors.append(f"Page quasiment vide ({len(words)} mot(s) detecte(s)) — rendu probablement casse.")

            # Detecte deux mots dont les rectangles se chevauchent nettement
            # (hors mots naturellement adjacents sur la meme ligne) : signe
            # d'un element de mise en page qui en recouvre un autre.
            def _iou(a, b):
                ix0, iy0 = max(a['x0'], b['x0']), max(a['top'], b['top'])
                ix1, iy1 = min(a['x1'], b['x1']), min(a['bottom'], b['bottom'])
                if ix1 <= ix0 or iy1 <= iy0:
                    return 0.0
                inter = (ix1 - ix0) * (iy1 - iy0)
                area_a = (a['x1'] - a['x0']) * (a['bottom'] - a['top'])
                area_b = (b['x1'] - b['x0']) * (b['bottom'] - b['top'])
                return inter / max(1.0, min(area_a, area_b))

            overlaps = []
            for i in range(len(words)):
                for j in range(i + 1, len(words)):
                    wa, wb = words[i], words[j]
                    same_line = abs(wa['top'] - wb['top']) < 2
                    if same_line:
                        continue  # mots cote a cote sur une meme ligne : normal
                    if _iou(wa, wb) > 0.3:
                        overlaps.append(f"'{wa['text']}' chevauche '{wb['text']}' (lignes differentes)")
            if overlaps:
                errors.append(f"{len(overlaps)} chevauchement(s) de texte detecte(s) : " + "; ".join(overlaps[:5]))
    except Exception as e:
        warnings.append(f"Verification visuelle du PDF impossible ({e}).")

    return {"ok": len(errors) == 0, "errors": errors, "warnings": warnings}


# ---------------------------------------------------------------------------
# Relecture visuelle par IA (Claude) - complement des controles
# deterministes ci-dessus. verify_podium_pdf() ne peut reperer que les
# categories de problemes anticipees a l'avance (texte qui se chevauche).
# Cette fonction va plus loin : elle montre une image du PDF genere a un
# modele Claude et lui demande une relecture libre, capable en principe de
# reperer un probleme visuel jamais vu auparavant (couleur illisible,
# element mal aligne, incoherence entre le chiffre affiche et le contexte,
# etc.) — exactement le type de controle fait manuellement tout au long de
# cette conversation. Degradation gracieuse si PyMuPDF ou le SDK Anthropic
# ne sont pas installes, ou si aucune cle API n'est configuree : la
# relecture est alors sautee (warning, jamais bloquant), le reste du
# pipeline (controles deterministes) continue de s'appliquer normalement.
# ---------------------------------------------------------------------------

try:
    import fitz  # PyMuPDF - rendu PDF->image sans dependance systeme (contrairement a poppler)
except ImportError:
    fitz = None

try:
    import anthropic as _anthropic_sdk
except ImportError:
    _anthropic_sdk = None

DEFAULT_VISION_MODEL = "claude-sonnet-5"

_VISION_PROMPT = """Tu relis le PDF hebdomadaire "Podium Performance Drive" d'un magasin \
Intermarché (charte rouge/noir/blanc), affiché en salle de pause et envoyé par email à un \
responsable. Ton rôle : repérer tout défaut visuel qui rendrait ce document gênant à \
diffuser tel quel, exactement comme le ferait un relecteur humain attentif avant envoi.

Vérifie en particulier :
- Chevauchement ou texte coupé/tronqué entre deux éléments (ex: le classement qui empiète \
sur le bloc "Taux de rupture", un nom qui déborde de sa carte).
- Zone blanche ou manifestement vide qui ne devrait pas l'être.
- Couleur de texte illisible sur son fond (ex: texte clair sur fond clair).
- Incohérence visible entre les chiffres affichés (ex: un podium où le rang 1 affiche une \
productivité plus faible que le rang 2).
- Tout élément visuellement cassé, mal aligné, ou qui a l'air d'un bug de génération.

Un texte simplement dense ou une mise en page serrée n'est PAS un défaut si tout reste \
lisible et rien ne se chevauche.

Réponds UNIQUEMENT avec un JSON strict, sans texte autour, au format :
{"ok": true ou false, "issues": ["description precise du probleme 1", ...]}
Si tout est correct, renvoie {"ok": true, "issues": []}."""


def render_pdf_page_png(pdf_path, page_index=0, zoom=2.0):
    """Rend une page du PDF en PNG (bytes) via PyMuPDF, sans dependance
    systeme externe (contrairement a pdftoppm/poppler). Retourne None si
    PyMuPDF n'est pas installe ou si le rendu echoue."""
    if fitz is None:
        return None
    try:
        doc = fitz.open(pdf_path)
        try:
            if page_index >= len(doc):
                return None
            page = doc[page_index]
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom))
            return pix.tobytes("png")
        finally:
            doc.close()
    except Exception:
        return None


def ai_visual_review(pdf_path, api_key=None, model=None):
    """Demande a un modele Claude de relire visuellement la premiere page du
    PDF genere. Retourne {"ok": bool, "issues": [...], "skipped": bool,
    "warnings": [...]}. `skipped=True` (avec ok=True) si la relecture n'a
    pas pu avoir lieu (dependance manquante ou cle API absente) — c'est un
    warning non-bloquant, pas un echec : le pipeline continue avec les
    seuls controles deterministes dans ce cas."""
    import base64, json as _json, os as _os

    api_key = api_key or _os.environ.get("ANTHROPIC_API_KEY")
    model = model or _os.environ.get("ANTHROPIC_VISION_MODEL", DEFAULT_VISION_MODEL)

    if fitz is None:
        return {"ok": True, "issues": [], "skipped": True,
                "warnings": ["PyMuPDF non installe : relecture visuelle IA sautee."]}
    if _anthropic_sdk is None:
        return {"ok": True, "issues": [], "skipped": True,
                "warnings": ["SDK anthropic non installe : relecture visuelle IA sautee."]}
    if not api_key:
        return {"ok": True, "issues": [], "skipped": True,
                "warnings": ["ANTHROPIC_API_KEY absente : relecture visuelle IA sautee."]}

    png_bytes = render_pdf_page_png(pdf_path)
    if not png_bytes:
        return {"ok": True, "issues": [], "skipped": True,
                "warnings": ["Rendu PNG du PDF impossible : relecture visuelle IA sautee."]}

    try:
        client = _anthropic_sdk.Anthropic(api_key=api_key)
        response = client.messages.create(
            model=model,
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {
                        "type": "base64", "media_type": "image/png",
                        "data": base64.b64encode(png_bytes).decode("ascii"),
                    }},
                    {"type": "text", "text": _VISION_PROMPT},
                ],
            }],
        )
        text = "".join(b.text for b in response.content if getattr(b, "type", None) == "text").strip()
        # Tolerant aux modeles qui entourent le JSON de ```...``` malgre la consigne.
        if text.startswith("```"):
            text = text.strip("`")
            text = text[4:] if text.lower().startswith("json") else text
        parsed = _json.loads(text)
        return {
            "ok": bool(parsed.get("ok", False)),
            "issues": list(parsed.get("issues", [])),
            "skipped": False,
            "warnings": [],
        }
    except Exception as e:
        # Une relecture IA qui echoue techniquement (timeout, reponse
        # malformee, quota...) ne doit jamais bloquer tout le pipeline
        # hebdomadaire a elle seule -> warning, pas erreur.
        return {"ok": True, "issues": [], "skipped": True,
                "warnings": [f"Relecture visuelle IA indisponible ({e})."]}
